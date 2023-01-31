import os
import sys
import openpyxl
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def setUpDriver():
    safari = webdriver.Safari()
    safari.maximize_window()
    safari.implicitly_wait(15)
    return safari

def waitForWebsiteLoadingComplete(driver):
    table = driver.find_element(By.ID, "hasResult")
    thead = table.find_element(By.CLASS_NAME, "cubinvest-l-table__tr")
    tbody = table.find_element(By.ID, "resultContainer")
    rows = tbody.find_elements(By.CLASS_NAME, "cubinvest-l-table__tr")
    lastRow = rows[len(rows) - 1]
    cells = lastRow.find_elements(By.XPATH, './/div[contains(@class, "cubinvest-l-table__td")][not(@style="display: none;")][not(@style="display:none;")]')
    nameCell = cells[2]
    etfId = nameCell.find_element(By.XPATH, './/div[@class="cubinvest-l-table__item cubinvest-l-table__item--fundId"]')
    WebDriverWait(driver, 15).until_not(EC.text_to_be_present_in_element((By.XPATH, '//*[@id="' + etfId.text + '-DjMktPrice"]'), "-"))
    return (table, thead, tbody)

def initialExcelFormat(workbook):
    sheet = workbook.worksheets[0]
    sheet.row_dimensions[1].height = 20
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 10
    return sheet

def writeTableHeadTitle(sheet, thead):
    columnTitles = thead.find_elements(By.XPATH, '//div[@class="cubinvest-l-table__th"][not(@style="display: none;")]')
    columnTitles.pop(0)
    columnTitles.pop(len(columnTitles)-1)
    for idx, title in enumerate(columnTitles):
        sheet.cell(row = 1, column = idx + 1).value = title.text.strip()
        sheet.cell(row = 1, column = idx + 1).font = Font(bold=True)

def writeTableData(sheet, tbody):
    rows = tbody.find_elements(By.CLASS_NAME, "cubinvest-l-table__tr")
    for RowIdx, row in enumerate(rows):
        cells = row.find_elements(By.XPATH, './/div[contains(@class, "cubinvest-l-table__td")][not(@style="display: none;")][not(@style="display:none;")]')
        cells.pop(0)
        cells.pop(len(cells)-1)
        for cellIdx, cell in enumerate(cells):
            if (cell.text.__contains__("基金")):
                sheet.cell(row=RowIdx+2, column=cellIdx+1).value = cell.text[:cell.text.index("基金")+2] + "\n" + cell.text[cell.text.index("基金")+2:]
            elif (cell.get_attribute("data-th") == "DjMktPrice"):
                elements = cell.find_elements(By.CLASS_NAME, 'cubinvest-l-table__item')
                sheet.cell(row=RowIdx+2, column=cellIdx+1).value = elements[0].text + "\n" + elements[1].text
            else:
                sheet.cell(row=RowIdx+2, column=cellIdx+1).value = cell.text
                sheet.cell(row=RowIdx+2, column=cellIdx+1).font = Font(color='ff0000') if cell.text.startswith("-") else Font(color='000000')

try:
    safari = setUpDriver()
    safari.get("https://www.cathaybk.com.tw/cathaybk/personal/investment/etf/search/?hotissue=05")

    if os.path.isfile('Cathay-United-Bank-ETF.xlsx'): # open or new a xlsx file
        workbook = openpyxl.load_workbook('Cathay-United-Bank-ETF.xlsx')
    else:
        workbook = openpyxl.Workbook()

    sheet = initialExcelFormat(workbook)
    (table, thead, tbody) = waitForWebsiteLoadingComplete(safari)
    writeTableHeadTitle(sheet, thead)
    writeTableData(sheet, tbody)

    workbook.save('Cathay-United-Bank-ETF.xlsx')
    workbook.close()
    safari.quit()
except Exception:
    if 'workbook' in vars():
        workbook.close()
    if 'safari' in vars():
        safari.quit()
    os.execv(sys.executable, ['python3'] + sys.argv)
